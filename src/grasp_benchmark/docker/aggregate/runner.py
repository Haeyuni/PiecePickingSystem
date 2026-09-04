import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw

METHODS = ('PCA_Normal', 'GG-CNN', 'GraspNet_baseline', 'Contact_GraspNet')
COLUMNS = ('scene_id', 'method', 'status', 'candidate_count', 'valid_width_count', 'best_score',
           'initialization_ms', 'inference_ms', 'width_m', 'x_m', 'y_m', 'z_m', 'width_validation',
           'candidate_definition', 'note', 'error_message', 'tested_at')
COLORS = {'PCA_Normal': '#e74c3c', 'GG-CNN': '#2ecc71', 'GraspNet_baseline': '#3498db', 'Contact_GraspNet': '#f1c40f'}


def not_run(scene_id, method):
    return {'scene_id': scene_id, 'method': method, 'status': 'NOT_RUN', 'candidate_count': 0,
            'valid_width_count': 0, 'best_score': None, 'initialization_ms': None, 'inference_ms': None,
            'width_m': None, 'x_m': None, 'y_m': None, 'z_m': None, 'width_validation': 'NOT_CHECKED',
            'candidate_definition': '', 'note': '', 'error_message': '컨테이너가 결과 JSON을 만들지 못함', 'tested_at': ''}


def preview(scene_path, rows, output):
    raw = np.load(scene_path, allow_pickle=False)
    image = Image.fromarray(np.asarray(raw['rgb']).astype(np.uint8)).convert('RGB')
    draw = ImageDraw.Draw(image)
    K = np.asarray(raw['K'], dtype=np.float32).reshape(3, 3)
    for row in rows:
        if row['status'] != 'OK' or row.get('z_m') in (None, 0):
            continue
        u = row.get('u_px')
        v = row.get('v_px')
        if u is None or v is None:
            u = int(round(row['x_m'] * K[0, 0] / row['z_m'] + K[0, 2]))
            v = int(round(row['y_m'] * K[1, 1] / row['z_m'] + K[1, 2]))
        color = COLORS[row['method']]
        draw.ellipse((u - 7, v - 7, u + 7, v + 7), outline=color, width=3)
        draw.text((u + 10, v - 8), row['method'], fill=color)
    image.save(output)


def main(scene_file):
    scene_id = Path(scene_file).stem
    results = Path('/results')
    rows = []
    for method in METHODS:
        name = method.lower().replace('-', '_')
        path = results / f'{scene_id}_{name}.json'
        rows.append(json.loads(path.read_text()) if path.is_file() else not_run(scene_id, method))
    dataframe = pd.DataFrame(rows).reindex(columns=COLUMNS)
    dataframe.to_csv(results / f'{scene_id}_comparison.csv', index=False)
    summary = (dataframe.groupby('method', dropna=False)
               .agg(run_count=('scene_id', 'count'), ok_count=('status', lambda x: int((x == 'OK').sum())),
                    valid_width_count=('valid_width_count', 'sum'), inference_ms=('inference_ms', 'mean'))
               .reset_index().round(2))
    notes = pd.DataFrame({'note': [
        'best_score는 모델별 내부 점수로 스케일과 의미가 다르므로 모델 간 우열 비교에 사용하지 않습니다.',
        'candidate_count는 모델별 생성 단계가 달라 후보 수 자체로 모델 순위를 정할 수 없습니다.',
        'Contact-GraspNet 공개 출력에는 RG2 그리퍼 폭이 없어 width_validation=UNAVAILABLE이 정상입니다.',
        '이 결과는 오프라인 후보 비교입니다. 충돌 검사, IK, 실제 RG2 파지 성공은 포함하지 않습니다.',
    ]})
    output = results / f'{scene_id}_comparison.xlsx'
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name='offline_results', index=False)
        summary.to_excel(writer, sheet_name='summary', index=False)
        notes.to_excel(writer, sheet_name='notes', index=False)
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = 'A2'
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='2E6EB5')
            for index, column in enumerate(worksheet.iter_cols(), 1):
                worksheet.column_dimensions[get_column_letter(index)].width = min(max(len(str(cell.value or '')) for cell in column) + 2, 55)
    preview(scene_file, rows, results / f'{scene_id}_comparison_preview.png')


if __name__ == '__main__':
    main(sys.argv[1])
