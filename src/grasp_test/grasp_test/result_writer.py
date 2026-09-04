import csv
import json
from pathlib import Path

FIELDS = ('scene_id', 'round_id', 'trial_id', 'method', 'status', 'candidate_count', 'valid_width_count',
          'model_init_ms', 'inference_ms', 'total_elapsed_ms', 'candidate_pose_camera', 'candidate_pose_robot',
          'transform_ok', 'ik_ok', 'rg2_grip_state', 'reobservation_ok', 'pick_success', 'failure_code', 'note')


class ResultWriter:
    def __init__(self, directory):
        self.directory = Path(directory)
        self.directory.mkdir(parents=True, exist_ok=True)
        self.rows = []

    def add(self, row):
        row = {field: row.get(field, '') for field in FIELDS}
        self.rows.append(row)
        (self.directory / f"{row['trial_id']}.json").write_text(json.dumps(row, ensure_ascii=False, indent=2))

    def save(self):
        csv_path = self.directory / 'live_physical_comparison.csv'
        with csv_path.open('w', newline='', encoding='utf-8') as stream:
            writer = csv.DictWriter(stream, fieldnames=FIELDS)
            writer.writeheader(); writer.writerows(self.rows)
        try:
            import openpyxl
            workbook = openpyxl.Workbook(); sheet = workbook.active; sheet.title = 'trials'
            sheet.append(FIELDS)
            for row in self.rows:
                sheet.append([row[field] for field in FIELDS])
            summary = workbook.create_sheet('summary')
            summary.append(['method', 'trials', 'success_count', 'success_rate', 'avg_inference_ms', 'avg_total_ms', 'failure_codes'])
            for method in sorted({row['method'] for row in self.rows}):
                rows = [row for row in self.rows if row['method'] == method]
                success = sum(row['pick_success'] is True for row in rows)
                numbers = lambda field: [float(row[field]) for row in rows if row[field] not in ('', None)]
                failures = sorted({row['failure_code'] for row in rows if row['failure_code']})
                summary.append([method, len(rows), success, success / len(rows) if rows else 0, *(sum(v) / len(v) if (v := numbers(f)) else None for f in ('inference_ms', 'total_elapsed_ms')), ', '.join(failures)])
            workbook.create_sheet('notes').append(['best_score is model-internal and excluded. Select by real pick success rate, failure types, and total elapsed time.'])
            workbook.save(self.directory / 'live_physical_comparison.xlsx')
        except ImportError:
            raise RuntimeError('openpyxl이 없어 XLSX를 만들 수 없습니다. .venv에 openpyxl을 설치하세요.')
